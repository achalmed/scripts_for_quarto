"""
lib/excel_writer.py
===================
Crea y actualiza el archivo Excel de metadatos.

Responsabilidades:
  - Generar la hoja METADATOS con encabezados estilizados.
  - Rellenar filas a partir del YAML de cada artículo.
  - Modo incremental: agregar sólo artículos nuevos preservando el resto.
  - Agregar columnas nuevas a un Excel existente.
  - Generar la hoja INSTRUCCIONES.

Depende de: config, yaml_parser, field_mapper, collector.
"""

from pathlib import Path
from typing import Dict, List, Optional, Set

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import ALL_FIELDS, VERSION
from .field_mapper import extract_value
from .yaml_parser import extract_yaml_only_index


# =============================================================================
# ESTILOS DE CABECERA
# =============================================================================

_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_cell(cell):
    cell.font      = _HEADER_FONT
    cell.fill      = _HEADER_FILL
    cell.alignment = _HEADER_ALIGN


# =============================================================================
# APERTURA LECTURA/ESCRITURA
# =============================================================================

def open_metadata_sheets(excel_path):
    """
    Abre la hoja METADATOS en dos vistas del mismo archivo:
      ws        → para escribir (preserva las fórmulas del usuario)
      ws_values → para leer valores CALCULADOS (data_only) en vez del
                  texto de la fórmula (p.ej. blog_nombre derivado)

    Devuelve (wb, ws, ws_values). Lanza la excepción original si falla.
    """
    wb = load_workbook(excel_path)
    ws = wb["METADATOS"]
    ws_values = load_workbook(excel_path, data_only=True)["METADATOS"]
    return wb, ws, ws_values


# =============================================================================
# RELLENO DE FILAS
# =============================================================================

def _fill_row(ws, row_idx: int, yaml_data: Dict, columns: List[str]):
    """Rellena una fila del Excel a partir de un dict YAML."""
    for col_idx, col_name in enumerate(columns, 1):
        try:
            value = extract_value(yaml_data, col_name)
            if value is not None:
                ws.cell(row_idx, col_idx, value)
        except Exception:
            pass


def _adjust_column_widths(ws):
    """Ajusta el ancho de cada columna según su contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 15), 60)


# =============================================================================
# CREACIÓN DE LA HOJA METADATOS
# =============================================================================

def build_metadata_sheet(
    wb: Workbook,
    df_files: pd.DataFrame,
    base_path: Path,
    columns: List[str] = ALL_FIELDS,
) -> None:
    """
    Construye la hoja METADATOS desde cero en el workbook dado.
    """
    ws = wb.create_sheet("METADATOS")

    # Encabezados
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, col_name)
        _style_header_cell(cell)

    # Datos
    print("\n📝 Extrayendo metadatos de cada artículo...\n")
    total = len(df_files)
    for row_idx, (_, row_data) in enumerate(df_files.iterrows(), 2):
        ws.cell(row_idx, 1, row_data["ruta_archivo"])
        ws.cell(row_idx, 2, row_data["blog_nombre"])
        ws.cell(row_idx, 3, row_data["tipo_documento"])

        file_path = base_path / row_data["ruta_archivo"]
        yaml_data = extract_yaml_only_index(file_path)
        if yaml_data:
            _fill_row(ws, row_idx, yaml_data, columns)

        if (row_idx - 1) % 10 == 0 or row_idx - 1 == total:
            print(f"  ✅ Procesados: {row_idx - 1}/{total} artículos")

    print(f"  ✅ Procesados: {total}/{total} artículos (100%)\n")

    _adjust_column_widths(ws)
    ws.freeze_panes = "A2"


# =============================================================================
# MODO INCREMENTAL
# =============================================================================

def append_new_articles(
    output_path: Path,
    df_files: pd.DataFrame,
    base_path: Path,
    columns: List[str] = ALL_FIELDS,
) -> bool:
    """
    Agrega solo los artículos que aún no existen en el Excel.
    Preserva fórmulas y formatos de las filas existentes.
    Devuelve True si agregó nuevos artículos, False si ya estaba al día.
    """
    try:
        wb = load_workbook(output_path)
        ws = wb["METADATOS"]

        existing: Set[str] = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing.add(row[0])

        df_new = df_files[~df_files["ruta_archivo"].isin(existing)]

        print(f"📊 Artículos en Excel existente: {len(existing)}")
        print(f"📊 Artículos encontrados ahora:  {len(df_files)}")

        if df_new.empty:
            print("\n✅ No hay artículos nuevos para agregar")
            print("   El Excel está actualizado\n")
            return False

        print(f"➕ Artículos nuevos a agregar: {len(df_new)}\n")

        last_row = ws.max_row
        print("📝 Agregando artículos nuevos...\n")
        for idx, (_, row_data) in enumerate(df_new.iterrows(), last_row + 1):
            ws.cell(idx, 1, row_data["ruta_archivo"])
            ws.cell(idx, 2, row_data["blog_nombre"])
            ws.cell(idx, 3, row_data["tipo_documento"])

            file_path = base_path / row_data["ruta_archivo"]
            yaml_data = extract_yaml_only_index(file_path)
            if yaml_data:
                _fill_row(ws, idx, yaml_data, columns)

            if (idx - last_row) % 10 == 0:
                print(f"  ✅ Procesados: {idx - last_row}/{len(df_new)}")

        print(f"  ✅ Procesados: {len(df_new)}/{len(df_new)} artículos (100%)\n")

        wb.save(output_path)
        total = len(existing) + len(df_new)
        print(f"✅ Excel actualizado (modo incremental): {output_path}")
        print(f"📊 Total artículos ahora: {total}")
        print(f"➕ Artículos nuevos agregados: {len(df_new)}")
        print(f"\n💡 Las fórmulas y formatos existentes se preservaron\n")
        return True

    except Exception as e:
        print(f"⚠️  Error en modo incremental: {e}")
        print("   Creando Excel nuevo en su lugar...\n")
        return False


# =============================================================================
# AGREGAR COLUMNAS NUEVAS
# =============================================================================

def add_columns_to_excel(
    excel_path: str,
    new_fields: List[str],
    base_path: Path,
    dry_run: bool = False,
):
    """
    Agrega nuevas columnas a la hoja METADATOS de un Excel existente y,
    si no es dry-run, llena los valores ya presentes en los archivos .qmd.
    """
    print(f"\n{'🔍 SIMULACIÓN' if dry_run else '➕ AGREGANDO'} COLUMNAS AL EXCEL\n")
    print("=" * 70)

    if not new_fields:
        print("⚠️  No hay campos nuevos para agregar")
        return

    try:
        wb = load_workbook(excel_path)
        ws = wb["METADATOS"]
        last_col = ws.max_column

        print(f"📊 Excel actual: {last_col} columnas")
        print(f"➕ Campos a agregar: {len(new_fields)}\n")

        for i, field in enumerate(new_fields, 1):
            new_col = last_col + i
            cell = ws.cell(1, new_col, field)
            _style_header_cell(cell)
            print(f"   {i}. {field} (columna {new_col})")

            if not dry_run:
                for row_idx in range(2, ws.max_row + 1):
                    ruta = ws.cell(row_idx, 1).value
                    if not ruta:
                        continue
                    file_path = base_path / ruta
                    if not file_path.exists():
                        continue
                    yaml_data = extract_yaml_only_index(file_path)
                    if yaml_data:
                        value = extract_value(yaml_data, field)
                        if value is not None:
                            ws.cell(row_idx, new_col, value)

        if not dry_run:
            wb.save(excel_path)
            print(f"\n✅ Excel actualizado: {excel_path}")
            print(f"📊 Total columnas ahora: {ws.max_column}")
        else:
            print(f"\n🔍 Simulación completada (no se guardaron cambios)")

        print("=" * 70 + "\n")

    except Exception as e:
        print(f"❌ Error: {e}")


# =============================================================================
# HOJA DE INSTRUCCIONES
# =============================================================================

def build_instructions_sheet(wb: Workbook):
    """Crea la hoja INSTRUCCIONES con guías de uso y formato."""
    ws = wb.create_sheet("INSTRUCCIONES", 0)

    lines = [
        [f"🎯 GUIA RAPIDA - SISTEMA DE GESTION DE METADATOS v{VERSION}"],
        [""],
        ["✅ NOVEDADES v2.0"],
        ["   • Arquitectura modular (fácil de extender)"],
        ["   • Actualización de blogs al formato pub_* actual"],
        ["   • Todos los comandos originales preservados"],
        ["   • Mejor manejo de errores y mensajes de progreso"],
        [""],
        ["📋 INSTRUCCIONES GENERALES"],
        [""],
        ["⚠️  COLUMNAS DE SOLO LECTURA (NO MODIFICAR):"],
        ["   • ruta_archivo: Ubicacion del archivo"],
        ["   • blog_nombre: Nombre del blog (pub_axiomata, website-achalma, etc.)"],
        ["   • tipo_documento: Tipo (stu/man/jou/doc)"],
        [""],
        ["✏️  COLUMNAS EDITABLES (Modificar libremente):"],
        ["   • Todas las demas columnas pueden ser editadas"],
        [""],
        ["🔄 FLUJO DE TRABAJO TIPICO:"],
        ["   1. Crear base de datos:"],
        ["      python main.py create-template ~/Documents"],
        ["   2. Abrir este Excel y editar metadatos"],
        ["   3. Guardar el archivo"],
        ["   4. Simular cambios (recomendado):"],
        ["      python main.py update ~/Documents quarto_metadata.xlsx --dry-run"],
        ["   5. Aplicar cambios:"],
        ["      python main.py update ~/Documents quarto_metadata.xlsx"],
        [""],
        ["=" * 72],
        [""],
        ["📝 FORMATO DE CAMPOS"],
        [""],
        ["✅ Campos TRUE/FALSE (booleanos):"],
        ["   • Escribir: TRUE o FALSE (MAYUSCULAS)"],
        ["   • draft = FALSE → artículo publicado (visible)"],
        ["   • draft = TRUE  → artículo borrador (oculto)"],
        ["   • eval  = TRUE  → evalúa bloques de código"],
        [""],
        ["📚 Campos de lista (separados por comas):"],
        ["   • keywords:   economia, estadistica, analisis"],
        ["   • tags:       python, tutorial, datos"],
        ["   • categories: Economia, Analisis Cuantitativo"],
        [""],
        ["📅 Fechas:"],
        ["   • Formato: MM/DD/YYYY"],
        ["   • Ejemplo: 12/19/2025"],
        [""],
        ["🔗 Links (JSON):"],
        ["   • links_enabled: TRUE o FALSE"],
        ["   • links_data: [{\"icon\":\"github\",\"url\":\"...\"}]"],
        [""],
        ["=" * 72],
        [""],
        ["📋 TIPOS DE DOCUMENTO"],
        [""],
        ["📌 STU (Estudiante): course, professor, duedate, note"],
        ["📌 JOU (Revista):    journal, volume, copyrightnotice, copyrightext"],
        ["📌 MAN (Manuscrito): floatsintext, numbered_lines, meta_analysis, mask"],
        ["📌 DOC (Documento):  floatsintext, numbered_lines"],
        [""],
        ["=" * 72],
        [""],
        ["👥 AUTORES (hasta 3)"],
        [""],
        ["Para cada autor N (1, 2 o 3):"],
        ["   • author_N_name:                Nombre completo"],
        ["   • author_N_corresponding:       TRUE (solo uno puede serlo)"],
        ["   • author_N_orcid:               0000-0001-6996-3364"],
        ["   • author_N_email:               correo@institucion.edu"],
        ["   • author_N_affiliation_name:    Universidad Nacional..."],
        ["   • author_N_affiliation_city:    Ayacucho"],
        ["   • author_N_roles:               conceptualization, writing"],
        [""],
        ["Roles CRediT válidos:"],
        ["   conceptualization, methodology, software, validation,"],
        ["   formal-analysis, investigation, resources, data-curation,"],
        ["   writing, visualization, supervision, project-administration,"],
        ["   funding-acquisition"],
        [""],
        ["=" * 72],
        [""],
        ["⚠️  PRECAUCIONES"],
        [""],
        ["🔒 Hacer BACKUP antes de actualizar (o usar Git)"],
        ["🔍 Probar con --dry-run antes de aplicar cambios reales"],
        ["✅ Booleanos: TRUE o FALSE (mayúsculas)"],
        ["📋 Listas: separar siempre con comas, sin corchetes"],
        ["💾 Guardar como .xlsx (no .xls ni .csv)"],
        [""],
        ["=" * 72],
        [""],
        ["🚀 COMANDOS UTILES"],
        [""],
        ["   python main.py create-config ~/Documents"],
        ["   python main.py create-template ~/Documents --config metadata_config.yml"],
        ["   python main.py create-template ~/Documents --incremental"],
        ["   python main.py update ~/Documents excel.xlsx --dry-run"],
        ["   python main.py update ~/Documents excel.xlsx --blog pub_axiomata"],
        ["   python main.py update ~/Documents excel.xlsx --filter-path 2025"],
        ["   python main.py detect-new-fields ~/Documents"],
        ["   python main.py find-differences ~/Documents excel.xlsx"],
        ["   python main.py sync-batch ~/Documents excel.xlsx --dry-run"],
        [""],
        ["🏷️  GESTION DE TAGS (destino: este Excel o el directorio de blogs)"],
        [""],
        ["   python main.py normalize-tags excel.xlsx --dry-run"],
        ["   python main.py replace-tags excel.xlsx viejo:nuevo otro:nuevo2"],
        ["   python main.py remove-tags excel.xlsx tag_obsoleto"],
        ["   python main.py add-tags excel.xlsx nuevo_tag --blog pub_axiomata"],
        ["   python main.py tag-stats ~/Documents --top 30"],
        ["   python main.py audit-tags ~/Documents"],
        [""],
        ["   Nota: sobre el Excel solo cambia la columna tags; luego"],
        ["   ejecutar 'update' para escribir los cambios en los .qmd"],
        [""],
        ["📅 SINCRONIZACION DESDE LA RUTA"],
        [""],
        ["   python main.py sync-dates excel.xlsx --dry-run"],
        ["   python main.py sync-pdf-urls excel.xlsx --dry-run"],
        ["   (date desde la carpeta YYYY-MM-DD; pdf-url desde la ruta real)"],
        [""],
        ["=" * 72],
        [""],
        ["📞 SOPORTE"],
        [""],
        [f"👤 Autor: {' Edison Achalma'}"],
        ["📧 Email: elmer.achalma.09@unsch.edu.pe"],
        ["📍 Ubicacion: Ayacucho, Peru"],
        [f"📌 Version: {VERSION}"],
        [""],
        ["=" * 72],
    ]

    for row_idx, line in enumerate(lines, 1):
        cell = ws.cell(row_idx, 1, line[0])
        text = line[0]
        if text.startswith("🎯") or text.startswith("===") or text.startswith("="):
            cell.font = Font(bold=True, size=14, color="1F4E78")
        elif any(text.startswith(e) for e in ["📋", "📝", "👥", "⚠️", "🚀", "💡", "📞", "✅"]):
            cell.font = Font(bold=True, size=12, color="366092")
        elif text.startswith("   •") or text.startswith("   "):
            cell.font = Font(size=10)

    ws.column_dimensions["A"].width = 90
