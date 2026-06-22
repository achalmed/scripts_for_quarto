"""
lib/sync.py
===========
Sincronización bidireccional entre index.qmd y Excel.

Comandos cubiertos:
  find-differences  → detecta artículos con valores distintos
  sync-article      → sincroniza un artículo de forma interactiva
  sync-batch        → sincroniza múltiples artículos de forma interactiva

También contiene detect_new_fields para detectar campos YAML no
declarados en ALL_FIELDS.

Depende de: config, yaml_parser, field_mapper, qmd_updater, excel_writer.
"""

from pathlib import Path
from typing import Dict, List, Optional, Set

import pandas as pd

from .config import ALL_FIELDS
from .field_mapper import extract_value, apply_row_to_yaml
from .yaml_parser import extract_yaml_only_index, flatten_yaml_keys
from .qmd_updater import update_single_qmd


# =============================================================================
# COMPARACIÓN DE UN ARTÍCULO
# =============================================================================

def compare_article(file_path: Path, excel_row: pd.Series) -> Optional[Dict]:
    """
    Compara los metadatos de index.qmd contra los valores del Excel.

    Devuelve un dict con:
      ruta          : str
      differences   : {campo: {index_value, excel_value}}
      only_in_index : [campos]
      only_in_excel : [campos]
    O None si no se puede analizar el archivo.
    """
    yaml_data = extract_yaml_only_index(file_path)
    if yaml_data is None:
        return None

    differences = {}
    only_in_index = []
    only_in_excel = []

    skip_fields = {"ruta_archivo", "blog_nombre", "fecha_creacion"}

    for field in ALL_FIELDS:
        if field in skip_fields:
            continue

        idx_val = extract_value(yaml_data, field)

        exc_val = excel_row.get(field)
        if pd.isna(exc_val) if isinstance(exc_val, float) else False:
            exc_val = None
        if isinstance(exc_val, str) and exc_val.strip() == "":
            exc_val = None

        # Normalizar booleanos para comparar
        if isinstance(idx_val, bool):
            idx_val = "TRUE" if idx_val else "FALSE"
        if isinstance(exc_val, str) and exc_val.upper() in ("TRUE", "FALSE"):
            exc_val = exc_val.upper()

        if idx_val is not None and exc_val is not None:
            if str(idx_val) != str(exc_val):
                differences[field] = {"index_value": idx_val, "excel_value": exc_val}
        elif idx_val is not None:
            only_in_index.append(field)
        elif exc_val is not None:
            only_in_excel.append(field)

    return {
        "ruta": str(file_path),
        "differences": differences,
        "only_in_index": only_in_index,
        "only_in_excel": only_in_excel,
    }


# =============================================================================
# ENCONTRAR DIFERENCIAS EN LOTE
# =============================================================================

def find_differences(
    base_path: Path,
    excel_path: str,
    blog_filter: Optional[str] = None,
    path_filter: Optional[str] = None,
    max_show: int = 10,
) -> List[Dict]:
    """
    Encuentra todos los artículos con diferencias entre index.qmd y Excel.
    """
    print(f"\n🔍 BUSCANDO DIFERENCIAS ENTRE INDEX.QMD Y EXCEL\n")
    print("=" * 70)

    try:
        df = pd.read_excel(excel_path, sheet_name="METADATOS")
    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
        return []

    if blog_filter:
        df = df[df["blog_nombre"] == blog_filter]
        print(f"🔍 Filtro: blog = '{blog_filter}'")
    if path_filter:
        df = df[df["ruta_archivo"].str.contains(path_filter, case=False, na=False)]
        print(f"🔍 Filtro: ruta contiene '{path_filter}'")

    print(f"📊 Artículos a analizar: {len(df)}\n")

    articles_with_diff = []
    for _, row in df.iterrows():
        ruta = row.get("ruta_archivo")
        if pd.isna(ruta):
            continue
        file_path = base_path / ruta
        if not file_path.exists():
            continue
        comp = compare_article(file_path, row)
        if not comp:
            continue
        if comp["differences"] or comp["only_in_index"] or comp["only_in_excel"]:
            articles_with_diff.append(comp)

    print("=" * 70)
    print(f"\n📊 RESUMEN:")
    print(f"   Total analizados:          {len(df)}")
    print(f"   Artículos con diferencias: {len(articles_with_diff)}")
    print(f"   Artículos sincronizados:   {len(df) - len(articles_with_diff)}")

    if articles_with_diff:
        show = min(max_show, len(articles_with_diff))
        print(f"\n📋 MOSTRANDO {show} DE {len(articles_with_diff)} ARTÍCULOS CON DIFERENCIAS:\n")
        for i, art in enumerate(articles_with_diff[:show], 1):
            print(f"\n{i}. 📄 {art['ruta']}")
            if art["differences"]:
                print(f"   ⚠️  Diferencias ({len(art['differences'])} campos):")
                for field, diff in list(art["differences"].items())[:5]:
                    print(f"      • {field}:")
                    print(f"         index.qmd: {str(diff['index_value'])[:40]}")
                    print(f"         Excel:     {str(diff['excel_value'])[:40]}")
                extra = len(art["differences"]) - 5
                if extra > 0:
                    print(f"      ... y {extra} diferencias más")
            if art["only_in_index"]:
                print(f"   📝 Solo en index.qmd: {', '.join(art['only_in_index'][:5])}")
            if art["only_in_excel"]:
                print(f"   📊 Solo en Excel:     {', '.join(art['only_in_excel'][:5])}")

        remaining = len(articles_with_diff) - show
        if remaining > 0:
            print(f"\n... y {remaining} artículos más con diferencias")

    print("\n" + "=" * 70 + "\n")
    return articles_with_diff


# =============================================================================
# SINCRONIZACIÓN INDIVIDUAL INTERACTIVA
# =============================================================================

def sync_single_interactive(
    base_path: Path,
    file_path: Path,
    excel_path: str,
    dry_run: bool = False,
):
    """Sincroniza un solo artículo, preguntando la dirección al usuario."""
    try:
        df = pd.read_excel(excel_path, sheet_name="METADATOS")
    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
        return

    ruta_rel = str(file_path.relative_to(base_path))
    matches = df[df["ruta_archivo"] == ruta_rel]
    if matches.empty:
        print(f"❌ Artículo no encontrado en Excel: {ruta_rel}")
        return

    row = matches.iloc[0]
    comp = compare_article(file_path, row)
    if not comp:
        print(f"⚠️  No se pudo analizar: {ruta_rel}")
        return

    print(f"\n📄 ANÁLISIS: {ruta_rel}\n")
    print("=" * 70)

    total = (
        len(comp["differences"])
        + len(comp["only_in_index"])
        + len(comp["only_in_excel"])
    )

    if total == 0:
        print("✅ Artículo sincronizado (sin diferencias)")
        print("=" * 70 + "\n")
        return

    print(f"⚠️  DIFERENCIAS ENCONTRADAS: {total}\n")
    if comp["differences"]:
        print(f"📊 Campos con valores diferentes ({len(comp['differences'])}):\n")
        for field, diff in comp["differences"].items():
            print(f"   • {field}:")
            print(f"      index.qmd: {diff['index_value']}")
            print(f"      Excel:     {diff['excel_value']}")
    if comp["only_in_index"]:
        print(f"📝 Solo en index.qmd: {', '.join(comp['only_in_index'])}\n")
    if comp["only_in_excel"]:
        print(f"📊 Solo en Excel: {', '.join(comp['only_in_excel'])}\n")

    print("=" * 70)
    print("\n💡 ¿Qué desea hacer?\n")
    print("   1. Actualizar Excel desde index.qmd (index.qmd → Excel)")
    print("   2. Actualizar index.qmd desde Excel (Excel → index.qmd)")
    print("   3. Cancelar (no hacer nada)")

    choice = "1" if dry_run else input("\n👉 Seleccione opción (1/2/3): ").strip()

    if choice == "1":
        print(f"\n{'🔍 SIMULANDO' if dry_run else '✅ ACTUALIZANDO'} Excel...\n")
        _sync_index_to_excel(file_path, base_path, excel_path, dry_run)
    elif choice == "2":
        print(f"\n{'🔍 SIMULANDO' if dry_run else '✅ ACTUALIZANDO'} index.qmd...\n")
        update_single_qmd(file_path, row, dry_run, 1, 1)
    else:
        print("\n⏭️  Cancelado (sin cambios)")

    print()


def _sync_index_to_excel(
    file_path: Path,
    base_path: Path,
    excel_path: str,
    dry_run: bool,
):
    """
    Copia los valores del index.qmd a la fila correspondiente del Excel.
    """
    from openpyxl import load_workbook
    from .config import ALL_FIELDS

    yaml_data = extract_yaml_only_index(file_path)
    if not yaml_data:
        print("❌ No se pudo extraer YAML")
        return

    wb = load_workbook(excel_path)
    ws = wb["METADATOS"]
    ruta_rel = str(file_path.relative_to(base_path))

    row_idx = None
    for idx in range(2, ws.max_row + 1):
        if ws.cell(idx, 1).value == ruta_rel:
            row_idx = idx
            break

    if not row_idx:
        print("❌ Artículo no encontrado en Excel")
        return

    changes = []
    for col_idx, field in enumerate(ALL_FIELDS, 1):
        if field in ("ruta_archivo", "blog_nombre"):
            continue
        new_val = extract_value(yaml_data, field)
        old_val = ws.cell(row_idx, col_idx).value
        if new_val != old_val:
            if not dry_run:
                ws.cell(row_idx, col_idx, new_val)
            changes.append(f"{field}: {old_val} → {new_val}")

    if changes:
        print(f"📝 Cambios: {len(changes)}\n")
        for c in changes[:10]:
            print(f"   • {c}")
        if len(changes) > 10:
            print(f"   ... y {len(changes) - 10} más")
        if not dry_run:
            wb.save(excel_path)
            print(f"\n✅ Excel actualizado")
    else:
        print("ℹ️  Sin cambios (ya estaba sincronizado)")


# =============================================================================
# SINCRONIZACIÓN MASIVA INTERACTIVA (sync-batch)
# =============================================================================

def sync_batch_interactive(
    base_path: Path,
    excel_path: str,
    blog_filter: Optional[str] = None,
    path_filter: Optional[str] = None,
    dry_run: bool = False,
):
    """
    Muestra artículos con diferencias y pregunta cómo sincronizarlos.
    """
    articles = find_differences(
        base_path, excel_path, blog_filter, path_filter, max_show=20
    )

    if not articles:
        print("✅ Todos los artículos están sincronizados")
        return

    print(f"\n💡 OPCIONES DE SINCRONIZACIÓN MASIVA:\n")
    print("   1. Actualizar TODO desde index.qmd → Excel")
    print("   2. Actualizar TODO desde Excel → index.qmd")
    print("   3. Decidir UNO POR UNO")
    print("   4. Cancelar")

    choice = "1" if dry_run else input("\n👉 Seleccione opción (1/2/3/4): ").strip()

    if choice == "1":
        print(f"\n{'🔍 SIMULANDO' if dry_run else '✅ ACTUALIZANDO'} TODO → Excel\n")
        for art in articles:
            fp = Path(art["ruta"])
            print(f"📄 {fp.name}")
            _sync_index_to_excel(fp, base_path, excel_path, dry_run)

    elif choice == "2":
        print(f"\n{'🔍 SIMULANDO' if dry_run else '✅ ACTUALIZANDO'} TODO → index.qmd\n")
        try:
            df = pd.read_excel(excel_path, sheet_name="METADATOS")
        except Exception as e:
            print(f"❌ Error leyendo Excel: {e}")
            return
        for art in articles:
            fp = Path(art["ruta"])
            ruta_rel = str(fp.relative_to(base_path))
            row = df[df["ruta_archivo"] == ruta_rel].iloc[0]
            print(f"📄 {fp.name}")
            update_single_qmd(fp, row, dry_run, 1, 1)

    elif choice == "3":
        for i, art in enumerate(articles, 1):
            print(f"\n{'=' * 70}")
            print(f"ARTÍCULO {i}/{len(articles)}")
            print(f"{'=' * 70}")
            fp = Path(art["ruta"])
            sync_single_interactive(base_path, fp, excel_path, dry_run)
            if i < len(articles):
                cont = "s" if dry_run else input("\n👉 Continuar? (s/n): ").strip().lower()
                if cont != "s":
                    print("\n⏭️  Proceso cancelado")
                    break
    else:
        print("\n⏭️  Cancelado")


# =============================================================================
# DETECCIÓN DE CAMPOS NUEVOS
# =============================================================================

def detect_new_fields(
    base_path: Path,
    allowed_blogs: set,
    user_excluded_folders: set,
    verbose: bool = True,
) -> dict:
    """
    Detecta campos YAML en index.qmd que NO están en ALL_FIELDS.
    Útil para mantener la plantilla Excel actualizada.
    """
    from .collector import collect_index_files

    print("\n🔍 DETECCIÓN DE NUEVOS METADATOS\n")
    print("=" * 70)

    df_files = collect_index_files(
        base_path, allowed_blogs, user_excluded_folders, verbose=False
    )

    new_fields_by_file = {}
    all_new: Set[str] = set()

    for _, row in df_files.iterrows():
        fp = base_path / row["ruta_archivo"]
        yaml_data = extract_yaml_only_index(fp)
        if not yaml_data:
            continue
        flat = flatten_yaml_keys(yaml_data)
        new = flat - set(ALL_FIELDS)
        if new:
            new_fields_by_file[row["ruta_archivo"]] = new
            all_new.update(new)
            if verbose:
                print(f"\n📄 {row['ruta_archivo']}")
                print(f"   Nuevos campos: {', '.join(sorted(new))}")

    print("\n" + "=" * 70)
    print(f"📊 RESUMEN:")
    print(f"   Archivos con campos nuevos:    {len(new_fields_by_file)}")
    print(f"   Total campos nuevos únicos:    {len(all_new)}")

    if all_new:
        print(f"\n💡 Campos nuevos detectados:")
        for f in sorted(all_new):
            print(f"   • {f}")
        print(f"\n💡 Para agregar al Excel:")
        print(f"   python main.py add-columns ~/Documents excel.xlsx {' '.join(sorted(all_new))}")

    print("=" * 70 + "\n")
    return new_fields_by_file
