#!/usr/bin/env python3
"""
Sistema de Gestión de Metadatos para Blogs Quarto - Versión 1.2.1
Autor: Edison Achalma
Fecha: Diciembre 2025

Mejoras v1.2.1:
- FIX: copyrightnotice se convierte correctamente a int (no float)
- FIX: corresponding se convierte correctamente a bool (no float)
- FIX: tipo_documento se actualiza correctamente (crea documentmode cuando cambia)
- NUEVO: Modo incremental para create-template (preserva fórmulas de Excel)
- NUEVO: Solo agrega nuevos artículos sin tocar existentes

Mejoras v1.2:
- Excluye index.qmd que no son artículos (sin fecha en ruta)
- Una sola hoja "METADATOS" en Excel (sin separar por tipo)
- Prioriza datos de index.qmd sobre _metadata.yml
- Solo actualiza cuando hay diferencias
- Filtros avanzados para actualización selectiva
- Instrucciones con emojis intuitivos
- Procesamiento más detallado y verbose
"""

import os
import sys
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import yaml
from datetime import datetime
from typing import Dict, List, Optional, Any, Set
import re
import json


class QuartoMetadataManager:
    """Gestor principal de metadatos de blogs Quarto v1.2.1"""
    
    # Carpetas del sistema a excluir SIEMPRE
    SYSTEM_EXCLUDED_FOLDERS = {
        '_site', '_freeze', 'site_libs', '.git', '.quarto', 
        'node_modules', '__pycache__', '_extensions',
        '.venv', 'venv', 'env', 'assets', '_partials',
        'title-block-link-buttons', 'Excalidraw'
    }
    
    # Archivos index.qmd a excluir (configuración)
    EXCLUDED_INDEX_FILES = {
        '_contenido-inicio.qmd', '_contenido-final.qmd', 
        '_contenido_posts.qmd', '_contenido_economia-preuniversitaria.qmd',
        '_contenido_inteligencia-comercial.qmd', '_contenido_talk.qmd',
        '_contenido_teching.qmd', '404.qmd', 'contact.qmd', 
        'accessibility.qmd', 'license.qmd', '_index.md', 'index.md'
    }
    
    # Campos de metadatos completos
    ALL_FIELDS = [
        'ruta_archivo', 'blog_nombre', 'tipo_documento',
        'title', 'shorttitle', 'subtitle', 
        'date', 'draft', 'abstract', 'description',
        'keywords', 'tags', 'categories',
        'image', 'eval', 'bibliography',
        'citation_type', 'citation_author', 'citation_pdf_url',
        'links_enabled', 'links_data',
        # Campos específicos (se incluyen todos)
        'course', 'professor', 'duedate', 'note',
        'journal', 'volume', 'copyrightnotice', 'copyrightext',
        'floatsintext', 'numbered_lines', 'meta_analysis', 'mask',
        # Autores
        'author_1_name', 'author_1_corresponding', 'author_1_orcid', 'author_1_email',
        'author_1_affiliation_name', 'author_1_affiliation_department',
        'author_1_affiliation_city', 'author_1_affiliation_region',
        'author_1_affiliation_country', 'author_1_roles',
        'author_2_name', 'author_2_orcid', 'author_2_affiliation_name', 'author_2_roles',
        'author_3_name', 'author_3_orcid', 'author_3_affiliation_name', 'author_3_roles'
    ]
    
    def __init__(self, base_path: str, config_file: Optional[str] = None):
        """Inicializa el gestor de metadatos"""
        self.base_path = Path(base_path).expanduser()
        if not self.base_path.exists():
            raise ValueError(f"❌ La ruta base no existe: {base_path}")
        
        self.config = self._load_config(config_file)
        self.user_excluded_folders = set(self.config.get('excluded_folders', []))
        self.allowed_blogs = set(self.config.get('allowed_blogs', []))
        self.excel_output_dir = Path(self.config.get('excel_output_dir', '.')).expanduser()
        self.excel_output_dir.mkdir(parents=True, exist_ok=True)
    
    def _load_config(self, config_file: Optional[str]) -> Dict:
        """Carga archivo de configuración YAML"""
        if config_file and Path(config_file).exists():
            with open(config_file, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        return {}
    
    def should_exclude_folder(self, folder_path: Path) -> bool:
        """Determina si una carpeta debe excluirse"""
        parts = set(folder_path.parts)
        return bool(parts & (self.SYSTEM_EXCLUDED_FOLDERS | self.user_excluded_folders))
    
    def should_exclude_file(self, file_path: Path) -> bool:
        """Determina si un archivo debe excluirse"""
        return file_path.name in self.EXCLUDED_INDEX_FILES
    
    def is_article_index(self, file_path: Path) -> bool:
        """
        Determina si un index.qmd es un artículo/publicación válido.
        Criterio: La carpeta padre debe comenzar con una fecha (YYYY-MM-DD)
        """
        parent_dir = file_path.parent.name
        
        # Patrón de fecha: YYYY-MM-DD o variantes
        date_pattern = r'^\d{4}-\d{2}-\d{2}'
        
        if re.match(date_pattern, parent_dir):
            return True
        
        # Si está directamente en blog/index.qmd o similar, NO es artículo
        grandparent = file_path.parent.parent.name
        if parent_dir in ['blog', 'posts', 'talk', 'teching', 'publication', 
                         'about', 'beschikbaarheid', 'appointment']:
            return False
        
        return False
    
    def is_allowed_blog(self, blog_name: str) -> bool:
        """Verifica si un blog está permitido según configuración"""
        if not self.allowed_blogs:
            return True
        return blog_name in self.allowed_blogs
    
    def find_metadata_yml(self, qmd_path: Path) -> Optional[Path]:
        """Busca el archivo _metadata.yml más cercano"""
        current_dir = qmd_path.parent
        while current_dir >= self.base_path:
            metadata_file = current_dir / '_metadata.yml'
            if metadata_file.exists():
                return metadata_file
            current_dir = current_dir.parent
        return None
    
    def load_metadata_yml(self, metadata_path: Path) -> Dict:
        """Carga contenido de _metadata.yml"""
        try:
            with open(metadata_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            print(f"⚠️  Error leyendo {metadata_path}: {e}")
            return {}
    
    def extract_yaml_from_qmd(self, file_path: Path, use_metadata: bool = True) -> Optional[Dict]:
        """
        Extrae YAML con PRIORIDAD a index.qmd sobre _metadata.yml
        Para visualización completa, fusiona ambos con prioridad a index.qmd
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            yaml_match = re.match(r'^---\s*\n(.*?)\n---', content, re.DOTALL)
            if not yaml_match:
                return None
            
            yaml_content = yaml_match.group(1)
            index_yaml = yaml.safe_load(yaml_content) or {}
            
            if not use_metadata:
                return index_yaml
            
            # Para visualización: fusionar pero indicar origen
            metadata_path = self.find_metadata_yml(file_path)
            if not metadata_path:
                return index_yaml
            
            base_yaml = self.load_metadata_yml(metadata_path)
            
            # FUSIÓN CON PRIORIDAD: index.qmd sobrescribe _metadata.yml
            # Solo usar _metadata.yml para campos que NO están en index.qmd
            result = base_yaml.copy()
            
            # Actualizar con valores de index.qmd (estos tienen prioridad)
            for key, value in index_yaml.items():
                if value is not None:
                    result[key] = value
            
            return result
            
        except Exception as e:
            print(f"⚠️  Error extrayendo YAML de {file_path.name}: {e}")
            return None
    
    def extract_yaml_only_index(self, file_path: Path) -> Optional[Dict]:
        """
        Extrae YAML SOLO del index.qmd (sin fusionar con _metadata.yml)
        Usado para extraer a Excel solo lo que está explícitamente definido
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            yaml_match = re.match(r'^---\s*\n(.*?)\n---', content, re.DOTALL)
            if not yaml_match:
                return None
            
            yaml_content = yaml_match.group(1)
            return yaml.safe_load(yaml_content) or {}
            
        except Exception as e:
            print(f"⚠️  Error extrayendo YAML de {file_path.name}: {e}")
            return None
    
    def detect_document_mode(self, yaml_data: Dict, file_path: Path) -> str:
        """
        Detecta tipo de documento SOLO del index.qmd (sin _metadata.yml)
        
        Prioridad:
        1. documentmode explícito en index.qmd
        2. format.apaquarto-pdf.documentmode en index.qmd
        3. Inferir por campos (course→stu, journal→jou)
        4. Si no hay nada, retornar None (usará _metadata.yml)
        """
        # Extraer YAML solo del index.qmd (sin fusión)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            yaml_match = re.match(r'^---\s*\n(.*?)\n---', content, re.DOTALL)
            if not yaml_match:
                return None
            
            yaml_content = yaml_match.group(1)
            index_only_yaml = yaml.safe_load(yaml_content) or {}
            
            # 1. Buscar documentmode directo en index.qmd
            if 'documentmode' in index_only_yaml:
                mode = index_only_yaml['documentmode']
                if mode in ['stu', 'man', 'jou', 'doc']:
                    return mode
            
            # 2. Buscar en format.apaquarto-pdf.documentmode en index.qmd
            if 'format' in index_only_yaml:
                formats = index_only_yaml['format']
                if isinstance(formats, dict) and 'apaquarto-pdf' in formats:
                    apa_config = formats['apaquarto-pdf']
                    if isinstance(apa_config, dict) and 'documentmode' in apa_config:
                        mode = apa_config['documentmode']
                        if mode in ['stu', 'man', 'jou', 'doc']:
                            return mode
            
            # 3. Inferir por campos específicos SOLO en index.qmd
            if 'course' in index_only_yaml or 'professor' in index_only_yaml:
                return 'stu'
            elif 'journal' in index_only_yaml and 'volume' in index_only_yaml:
                return 'jou'
            elif 'meta-analysis' in index_only_yaml or 'meta_analysis' in index_only_yaml:
                return 'man'
            
            # 4. Si no hay nada en index.qmd, retornar None
            return None
            
        except Exception as e:
            return None
    
    def collect_index_files(self, blog_name: Optional[str] = None, 
                           verbose: bool = True) -> pd.DataFrame:
        """
        Recolecta archivos index.qmd CON FILTRO de artículos
        Solo incluye archivos que son artículos válidos (con fecha en carpeta)
        """
        index_files = []
        
        if blog_name:
            blog_path = self.base_path / blog_name
            if not blog_path.exists():
                print(f"⚠️  El blog '{blog_name}' no existe")
                return pd.DataFrame()
            blogs_to_process = [blog_path]
        else:
            all_dirs = [d for d in self.base_path.iterdir() 
                       if d.is_dir() and not d.name.startswith('.')]
            
            if self.allowed_blogs:
                blogs_to_process = [d for d in all_dirs if d.name in self.allowed_blogs]
            else:
                blogs_to_process = all_dirs
        
        total_found = 0
        total_articles = 0
        total_skipped = 0
        
        for blog_dir in blogs_to_process:
            if not self.is_allowed_blog(blog_dir.name):
                continue
            
            print(f"\n📂 Procesando blog: {blog_dir.name}")
            blog_found = 0
            blog_articles = 0
            blog_skipped = 0
            
            for root, dirs, files in os.walk(blog_dir):
                dirs[:] = [d for d in dirs if not self.should_exclude_folder(Path(root) / d)]
                
                for file in files:
                    if file == 'index.qmd':
                        file_path = Path(root) / file
                        total_found += 1
                        blog_found += 1
                        
                        # Excluir archivos especiales
                        if self.should_exclude_file(file_path):
                            if verbose:
                                print(f"  ⏭️  Omitido (config): {file_path.name}")
                            total_skipped += 1
                            blog_skipped += 1
                            continue
                        
                        # FILTRO: Solo artículos (con fecha en carpeta)
                        if not self.is_article_index(file_path):
                            if verbose:
                                rel = file_path.relative_to(self.base_path)
                                print(f"  ⏭️  Omitido (no es artículo): {rel}")
                            total_skipped += 1
                            blog_skipped += 1
                            continue
                        
                        # Extraer YAML
                        yaml_data = self.extract_yaml_from_qmd(file_path)
                        if not yaml_data:
                            if verbose:
                                print(f"  ⚠️  Sin YAML: {file_path.name}")
                            total_skipped += 1
                            blog_skipped += 1
                            continue
                        
                        # Detectar tipo de documento SOLO del index.qmd
                        doc_type = self.detect_document_mode(yaml_data, file_path)
                        if doc_type is None:
                            doc_type = 'jou'  # Default si no está definido en ningún lugar
                        
                        try:
                            creation_time = datetime.fromtimestamp(file_path.stat().st_ctime)
                        except:
                            creation_time = datetime.now()
                        
                        rel_path = file_path.relative_to(self.base_path)
                        
                        index_files.append({
                            'blog_nombre': blog_dir.name,
                            'ruta_archivo': str(rel_path),
                            'tipo_documento': doc_type,
                            'fecha_creacion': creation_time,
                            'titulo': yaml_data.get('title', ''),
                            'draft': yaml_data.get('draft', True)
                        })
                        
                        total_articles += 1
                        blog_articles += 1
                        
                        if verbose:
                            print(f"  ✅ Artículo: {file_path.parent.name}/{file_path.name}")
            
            print(f"  📊 Blog '{blog_dir.name}': {blog_articles} artículos, {blog_skipped} omitidos")
        
        print(f"\n{'='*70}")
        print(f"📊 RESUMEN DE RECOLECCIÓN:")
        print(f"  📁 Total archivos encontrados: {total_found}")
        print(f"  ✅ Artículos válidos: {total_articles}")
        print(f"  ⏭️  Omitidos: {total_skipped}")
        print(f"{'='*70}\n")
        
        df = pd.DataFrame(index_files)
        
        if not df.empty:
            df = df.sort_values(['blog_nombre', 'tipo_documento', 'fecha_creacion'], 
                              ascending=[True, True, False])
        
        return df
    
    def create_excel_template(self, output_filename: str, blog_name: Optional[str] = None,
                             incremental: bool = False):
        """
        Crea plantilla Excel con UNA SOLA HOJA de metadatos
        
        Modo incremental: Solo agrega nuevos artículos sin tocar existentes
        Esto preserva fórmulas y formatos personalizados en Excel
        """
        print("🔍 Recolectando archivos index.qmd...")
        print("   (Solo se incluirán artículos/publicaciones con fecha)\n")
        
        df_files = self.collect_index_files(blog_name, verbose=True)
        
        if df_files.empty:
            print("⚠️  No se encontraron artículos válidos")
            return
        
        output_path = self.excel_output_dir / output_filename
        
        # MODO INCREMENTAL: Solo agregar nuevos artículos
        if incremental and output_path.exists():
            print("\n🔄 MODO INCREMENTAL: Preservando datos existentes")
            print("   Solo se agregarán artículos nuevos\n")
            
            try:
                wb_existing = load_workbook(output_path)
                ws_existing = wb_existing['METADATOS']
                
                existing_paths = set()
                for row in ws_existing.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0]:
                        existing_paths.add(row[0])
                
                print(f"📊 Artículos en Excel existente: {len(existing_paths)}")
                print(f"📊 Artículos encontrados ahora: {len(df_files)}")
                
                df_new = df_files[~df_files['ruta_archivo'].isin(existing_paths)]
                
                if df_new.empty:
                    print("\n✅ No hay artículos nuevos para agregar")
                    print("   El Excel está actualizado\n")
                    return
                
                print(f"➕ Artículos nuevos a agregar: {len(df_new)}\n")
                
                last_row = ws_existing.max_row
                
                print("📝 Agregando artículos nuevos...\n")
                for idx, (_, row_data) in enumerate(df_new.iterrows(), last_row + 1):
                    ws_existing.cell(idx, 1, row_data['ruta_archivo'])
                    ws_existing.cell(idx, 2, row_data['blog_nombre'])
                    ws_existing.cell(idx, 3, row_data['tipo_documento'])
                    
                    file_path = self.base_path / row_data['ruta_archivo']
                    yaml_data = self.extract_yaml_only_index(file_path)
                    
                    if yaml_data:
                        self._fill_excel_row_from_yaml(ws_existing, idx, yaml_data, self.ALL_FIELDS)
                    
                    if idx % 10 == 0:
                        print(f"  ✅ Procesados: {idx - last_row}/{len(df_new)} artículos")
                
                print(f"  ✅ Procesados: {len(df_new)}/{len(df_new)} artículos (100%)\n")
                
                wb_existing.save(output_path)
                
                print(f"✅ Excel actualizado (modo incremental): {output_path}")
                print(f"📊 Total artículos ahora: {len(existing_paths) + len(df_new)}")
                print(f"➕ Artículos nuevos agregados: {len(df_new)}")
                print(f"\n💡 Las fórmulas y formatos existentes se preservaron\n")
                
                return
                
            except Exception as e:
                print(f"⚠️  Error en modo incremental: {e}")
                print("   Creando Excel nuevo en su lugar...\n")
        
        # MODO NORMAL: Crear Excel desde cero
        wb = Workbook()
        wb.remove(wb.active)
        
        # HOJA ÚNICA DE METADATOS
        ws = wb.create_sheet("METADATOS")
        
        columns = self.ALL_FIELDS
        
        # Encabezados con estilo
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(1, col_idx, col_name)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Datos
        print("\n📝 Extrayendo metadatos de cada artículo...\n")
        for row_idx, (_, row_data) in enumerate(df_files.iterrows(), 2):
            ws.cell(row_idx, 1, row_data['ruta_archivo'])
            ws.cell(row_idx, 2, row_data['blog_nombre'])
            ws.cell(row_idx, 3, row_data['tipo_documento'])
            
            file_path = self.base_path / row_data['ruta_archivo']
            
            # IMPORTANTE: Extraer SOLO lo que está en index.qmd (sin _metadata.yml)
            yaml_data = self.extract_yaml_only_index(file_path)
            
            if yaml_data:
                self._fill_excel_row_from_yaml(ws, row_idx, yaml_data, columns)
                
            # Progreso
            if row_idx % 10 == 0:
                print(f"  ✅ Procesados: {row_idx - 1}/{len(df_files)} artículos")
        
        print(f"  ✅ Procesados: {len(df_files)}/{len(df_files)} artículos (100%)\n")
        
        # Ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 60)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        
        # Hoja de instrucciones
        self._create_instructions_sheet(wb)
        
        wb.save(output_path)
        
        print(f"✅ Plantilla Excel creada: {output_path}")
        print(f"📊 Total de artículos: {len(df_files)}")
        print(f"📁 Hojas: METADATOS (todos los artículos), INSTRUCCIONES")
        print(f"\n💡 Próximos pasos:")
        print(f"   1. Abrir: {output_path}")
        print(f"   2. Editar metadatos en hoja METADATOS")
        print(f"   3. Guardar")
        print(f"   4. Actualizar: python quarto_metadata_manager.py update \\")
        print(f"      {self.base_path} {output_path}")
        print(f"\n💡 Tip: Usa '--incremental' para agregar solo artículos nuevos:")
        print(f"   python quarto_metadata_manager.py create-template ... --incremental\n")
    
    def _create_instructions_sheet(self, wb: Workbook):
        """Crea hoja de instrucciones con EMOJIS intuitivos"""
        ws = wb.create_sheet("INSTRUCCIONES", 0)
        
        instructions = [
            ["🎯 GUIA RAPIDA - SISTEMA DE GESTION DE METADATOS v1.2.1"],
            [""],
            ["✅ CAMBIOS v1.2.1"],
            ["   • FIX: copyrightnotice como entero (2025, no 2025.0)"],
            ["   • FIX: corresponding como booleano (TRUE, no 1.0)"],
            ["   • FIX: tipo_documento actualiza correctamente"],
            ["   • NUEVO: Modo incremental (preserva formulas)"],
            ["📋 INSTRUCCIONES GENERALES"],
            [""],
            ["⚠️  COLUMNAS DE SOLO LECTURA (NO MODIFICAR):"],
            ["   • ruta_archivo: Ubicacion del archivo"],
            ["   • blog_nombre: Nombre del blog"],
            ["   • tipo_documento: Tipo (STU/MAN/JOU/DOC)"],
            [""],
            ["✏️  COLUMNAS EDITABLES (Modificar libremente):"],
            ["   • Todas las demas columnas pueden ser editadas"],
            [""],
            ["➕ AGREGAR NUEVOS ARTICULOS:"],
            ["   1. Crear el articulo en tu blog"],
            ["   2. Ejecutar: python ... create-template [ruta]"],
            ["   3. Los nuevos articulos apareceran en el Excel"],
            [""],
            ["🔄 APLICAR CAMBIOS:"],
            ["   1. Editar metadatos en este Excel"],
            ["   2. Guardar el archivo"],
            ["   3. Ejecutar: python ... update [ruta] [excel]"],
            ["   4. Solo se actualizan los index.qmd que tienen diferencias"],
            [""],
            ["========================================================================"],
            [""],
            ["📝 FORMATO DE CAMPOS"],
            [""],
            ["✅ Campos TRUE/FALSE (booleanos):"],
            ["   • Escribir: TRUE o FALSE (MAYUSCULAS)"],
            ["   • Ejemplo: draft = FALSE (para publicar)"],
            ["   • Ejemplo: eval = TRUE (para evaluar codigo)"],
            [""],
            ["📚 Campos de lista (separados por comas):"],
            ["   • keywords: economia, estadistica, analisis"],
            ["   • tags: python, tutorial, datos"],
            ["   • categories: Economia, Analisis Cuantitativo"],
            [""],
            ["📅 Fechas:"],
            ["   • Formato: MM/DD/YYYY"],
            ["   • Ejemplo: 12/19/2025"],
            ["   • Ejemplo: 01/15/2024"],
            [""],
            ["🔗 Links (JSON):"],
            ["   • links_enabled: TRUE o FALSE"],
            ["   • links_data: [{'icon':'github','url':'...'}]"],
            [""],
            ["========================================================================"],
            [""],
            ["📋 CAMPOS OBLIGATORIOS (Todos los tipos)"],
            [""],
            ["📌 IDENTIFICACION:"],
            ["   • title: Titulo principal del articulo"],
            ["   • shorttitle: Titulo corto (max 50 caracteres)"],
            ["   • subtitle: Subtitulo (opcional)"],
            [""],
            ["📅 PUBLICACION:"],
            ["   • date: Fecha de publicacion (MM/DD/YYYY)"],
            ["   • draft: FALSE=publicado, TRUE=borrador"],
            [""],
            ["📝 DESCRIPCION:"],
            ["   • abstract: Resumen academico (max 250 palabras)"],
            ["   • description: Descripcion breve para SEO"],
            [""],
            ["🏷️  CLASIFICACION:"],
            ["   • keywords: 3-5 palabras clave (separadas por comas)"],
            ["   • tags: Etiquetas para busqueda"],
            ["   • categories: Categorias del contenido"],
            [""],
            ["🖼️  MEDIOS:"],
            ["   • image: Nombre del archivo (ej: featured.png)"],
            [""],
            ["💻 CODIGO:"],
            ["   • eval: TRUE/FALSE (evaluar bloques de codigo)"],
            [""],
            ["📖 CITACION:"],
            ["   • citation_type: article-journal, book, etc."],
            ["   • citation_author: Nombre del autor"],
            ["   • citation_pdf_url: URL del PDF"],
            [""],
            ["📚 BIBLIOGRAFIA:"],
            ["   • bibliography: referencias.bib"],
            [""],
            ["========================================================================"],
            [""],
            ["📑 CAMPOS ESPECIFICOS POR TIPO"],
            [""],
            ["🎓 MODO ESTUDIANTE (STU)"],
            ["   Para: Trabajos academicos, tareas, proyectos"],
            [""],
            ["   • course: Nombre del curso"],
            ["     Ejemplo: Metodologia de Investigacion (ECON 101)"],
            [""],
            ["   • professor: Nombre del profesor"],
            ["     Ejemplo: Dr. Edison Achalma"],
            [""],
            ["   • duedate: Fecha de entrega"],
            ["     Ejemplo: 12/25/2025"],
            [""],
            ["   • note: Notas adicionales"],
            ["     Ejemplo: Codigo: 2020123456, Seccion: A"],
            [""],
            ["📰 MODO REVISTA (JOU)"],
            ["   Para: Articulos publicados en revistas"],
            [""],
            ["   • journal: Nombre de la revista"],
            ["     Ejemplo: Revista Peruana de Economia"],
            [""],
            ["   • volume: Volumen, numero y paginas"],
            ["     Ejemplo: 2025, Vol. 7, No. 1, 1--25"],
            [""],
            ["   • copyrightnotice: Anio de copyright"],
            ["     Ejemplo: 2025"],
            [""],
            ["   • copyrightext: Texto completo de copyright"],
            ["     Ejemplo: Todos los derechos reservados"],
            [""],
            ["📄 MODO MANUSCRITO (MAN)"],
            ["   Para: Manuscritos para envio a revistas"],
            [""],
            ["   • floatsintext: TRUE/FALSE"],
            ["     TRUE = Figuras en el texto"],
            ["     FALSE = Figuras al final (estandar)"],
            [""],
            ["   • numbered_lines: TRUE/FALSE"],
            ["     TRUE = Numerar lineas (util para revision)"],
            [""],
            ["   • meta_analysis: TRUE/FALSE"],
            ["     TRUE = Incluye meta-analisis"],
            [""],
            ["   • mask: TRUE/FALSE"],
            ["     TRUE = Ocultar autores (revision ciega)"],
            [""],
            ["📝 MODO DOCUMENTO (DOC)"],
            ["   Para: Documentos generales, informes, ensayos"],
            [""],
            ["   • floatsintext: TRUE/FALSE"],
            ["   • numbered_lines: TRUE/FALSE"],
            [""],
            ["========================================================================"],
            [""],
            ["👥 AUTORES (Hasta 3 autores)"],
            [""],
            ["Para cada autor (N = 1, 2, 3):"],
            [""],
            ["   • author_N_name: Nombre completo"],
            ["     Ejemplo: Edison Achalma"],
            [""],
            ["   • author_N_corresponding: TRUE/FALSE"],
            ["     ⚠️  Solo UNO puede ser TRUE"],
            [""],
            ["   • author_N_orcid: ID ORCID"],
            ["     Formato: 0000-0002-XXXX-XXXX"],
            [""],
            ["   • author_N_email: Email de contacto"],
            [""],
            ["   • author_N_affiliation_name: Institucion"],
            ["   • author_N_affiliation_department: Departamento"],
            ["   • author_N_affiliation_city: Ciudad"],
            ["   • author_N_affiliation_region: Region"],
            ["   • author_N_affiliation_country: Pais"],
            [""],
            ["   • author_N_roles: Roles CRediT (separados por comas)"],
            ["     Ejemplo: conceptualization, writing, analysis"],
            [""],
            ["========================================================================"],
            [""],
            ["⚠️  PRECAUCIONES IMPORTANTES"],
            [""],
            ["🔒 HACER BACKUP:"],
            ["   • Siempre hacer copia de seguridad antes de actualizar"],
            ["   • Usar Git para control de versiones"],
            [""],
            ["🔍 PROBAR PRIMERO:"],
            ["   • Usar --dry-run para ver cambios sin aplicar"],
            ["   • Comando: python ... update [ruta] [excel] --dry-run"],
            [""],
            ["✅ VERIFICAR FORMATO:"],
            ["   • Booleanos: TRUE o FALSE (mayusculas)"],
            ["   • Listas: separar con comas"],
            ["   • Fechas: MM/DD/YYYY"],
            [""],
            ["💾 GUARDAR CORRECTAMENTE:"],
            ["   • Guardar como .xlsx (Excel 2007+)"],
            ["   • NO usar .xls ni .csv"],
            [""],
            ["========================================================================"],
            [""],
            ["🚀 COMANDOS UTILES"],
            [""],
            ["🔍 Ver cambios sin aplicar:"],
            ["   python quarto_metadata_manager.py update [ruta] [excel] --dry-run"],
            [""],
            ["📁 Actualizar solo un blog:"],
            ["   python ... update [ruta] [excel] --blog axiomata"],
            [""],
            ["🎯 Actualizar solo una ruta especifica:"],
            ["   python ... update [ruta] [excel] --filter-path blog/posts/2024"],
            [""],
            ["📊 Crear base de datos de un blog:"],
            ["   python ... create-template [ruta] --blog axiomata"],
            [""],
            ["🌐 Crear base de datos general:"],
            ["   python ... create-template [ruta] --config config.yml"],
            [""],
            ["========================================================================"],
            [""],
            ["💡 CONSEJOS Y TIPS"],
            [""],
            ["✨ Para publicar un articulo:"],
            ["   1. Buscar el articulo en Excel"],
            ["   2. Cambiar draft de TRUE a FALSE"],
            ["   3. Guardar Excel"],
            ["   4. Ejecutar update"],
            [""],
            ["📝 Para cambiar tipo de documento:"],
            ["   1. Cambiar tipo_documento (STU/MAN/JOU/DOC)"],
            ["   2. Llenar campos especificos del tipo"],
            ["   3. Guardar y ejecutar update"],
            [""],
            ["🏷️  Para agregar tags nuevos:"],
            ["   1. Editar columna tags"],
            ["   2. Separar con comas"],
            ["   3. No usar punto y coma ni corchetes"],
            [""],
            ["========================================================================"],
            [""],
            ["📞 SOPORTE Y AYUDA"],
            [""],
            ["👤 Autor: Edison Achalma"],
            ["📧 Email: achalmaedison@gmail.com"],
            ["📍 Ubicacion: Ayacucho, Peru"],
            ["📌 Version: 1.2.0"],
            [""],
            ["🐛 Reportar problemas:"],
            ["   1. Ejecutar con --dry-run para diagnosticar"],
            ["   2. Verificar formato de datos"],
            ["   3. Revisar mensajes de error"],
            ["   4. Contactar con detalles del error"],
            [""],
            ["========================================================================"],
            [""],
            ["✅ CAMBIOS v1.2.0"],
            [""],
            ["🎯 Mejoras:"],
            ["   • Solo procesa articulos (con fecha en carpeta)"],
            ["   • Una sola hoja METADATOS (todos juntos)"],
            ["   • Prioriza index.qmd sobre _metadata.yml"],
            ["   • Solo actualiza cuando hay diferencias"],
            ["   • Filtros avanzados para actualizacion"],
            ["   • Instrucciones con emojis intuitivos"],
            ["   • Procesamiento mas detallado"],
            [""],
            ["========================================================================"],
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            cell = ws.cell(row_idx, 1, instruction[0])
            
            # Estilos según contenido
            text = instruction[0]
            if text.startswith("🎯") or text.startswith("==="):
                cell.font = Font(bold=True, size=14, color='1F4E78')
            elif any(text.startswith(emoji) for emoji in ["📋", "📝", "👥", "⚠️", "🚀", "💡", "📞", "✅"]):
                cell.font = Font(bold=True, size=12, color='366092')
            elif text.startswith("   •"):
                cell.font = Font(size=10)
        
        ws.column_dimensions['A'].width = 90
    
    def _fill_excel_row_from_yaml(self, ws, row_idx: int, yaml_data: Dict, columns: List[str]):
        """Llena una fila de Excel con datos extraídos del YAML"""
        for col_idx, col_name in enumerate(columns, 1):
            try:
                value = self._extract_yaml_value(yaml_data, col_name)
                if value is not None:
                    ws.cell(row_idx, col_idx, value)
            except Exception as e:
                continue
    
    def _extract_yaml_value(self, yaml_data: Dict, field_name: str) -> Any:
        """Extrae un valor específico del YAML"""
        simple_mapping = {
            'title': 'title', 'shorttitle': 'shorttitle', 'subtitle': 'subtitle',
            'date': 'date', 'draft': 'draft', 'abstract': 'abstract',
            'description': 'description', 'image': 'image', 'eval': 'eval',
            'bibliography': 'bibliography', 'course': 'course', 'professor': 'professor',
            'duedate': 'duedate', 'note': 'note', 'journal': 'journal',
            'volume': 'volume', 'copyrightnotice': 'copyrightnotice',
            'copyrightext': 'copyrightext', 'floatsintext': 'floatsintext',
            'numbered_lines': 'numbered-lines', 'meta_analysis': 'meta-analysis',
            'mask': 'mask'
        }
        
        if field_name in simple_mapping:
            value = yaml_data.get(simple_mapping[field_name])
            
            # FIX: Convertir copyrightnotice a int
            if field_name == 'copyrightnotice' and value is not None:
                try:
                    return int(value)
                except:
                    return value
            
            return value
        
        if field_name in ['keywords', 'tags', 'categories']:
            value = yaml_data.get(field_name, [])
            if isinstance(value, list):
                return ', '.join([str(v) for v in value])
            return value
        
        if field_name.startswith('citation_'):
            citation = yaml_data.get('citation')
            if not isinstance(citation, dict):
                return None
            
            if field_name == 'citation_type':
                return citation.get('type')
            elif field_name == 'citation_author':
                authors = citation.get('author', [])
                if isinstance(authors, list):
                    return ', '.join([str(a) for a in authors])
                return authors
            elif field_name == 'citation_pdf_url':
                return citation.get('pdf-url')
        
        if field_name == 'links_enabled':
            links = yaml_data.get('links')
            return links is not None and links != False
        elif field_name == 'links_data':
            links = yaml_data.get('links')
            if links and isinstance(links, (list, dict)):
                return json.dumps(links, ensure_ascii=False)
        
        if field_name.startswith('author_'):
            match = re.match(r'author_(\d+)_(.*)', field_name)
            if match:
                author_idx = int(match.group(1)) - 1
                field_suffix = match.group(2)
                
                authors = yaml_data.get('author', [])
                if not isinstance(authors, list) or author_idx >= len(authors):
                    return None
                
                author = authors[author_idx]
                
                if field_suffix == 'name':
                    return author.get('name')
                elif field_suffix == 'corresponding':
                    # FIX: Convertir corresponding a string TRUE/FALSE
                    corr = author.get('corresponding')
                    if corr is True:
                        return 'TRUE'
                    elif corr is False:
                        return 'FALSE'
                    return corr
                elif field_suffix == 'orcid':
                    return author.get('orcid')
                elif field_suffix == 'email':
                    return author.get('email')
                elif field_suffix == 'roles':
                    roles = author.get('role', [])
                    if isinstance(roles, list):
                        return ', '.join([str(r) for r in roles])
                    return roles
                elif field_suffix.startswith('affiliation_'):
                    aff_field = field_suffix.replace('affiliation_', '')
                    affiliations = author.get('affiliations', [])
                    if affiliations and isinstance(affiliations, list):
                        for aff in affiliations:
                            if isinstance(aff, dict):
                                return aff.get(aff_field)
        
        return None
    
    def update_yaml_from_excel(self, excel_path: str, 
                              blog_filter: Optional[str] = None,
                              path_filter: Optional[str] = None,
                              dry_run: bool = False):
        """
        Actualiza archivos desde Excel con FILTROS AVANZADOS
        
        Args:
            excel_path: Ruta del Excel
            blog_filter: Filtrar por blog específico
            path_filter: Filtrar por substring en ruta
            dry_run: Simulación sin aplicar cambios
        """
        print(f"\n📖 Leyendo Excel: {excel_path}\n")
        
        try:
            df = pd.read_excel(excel_path, sheet_name='METADATOS')
        except Exception as e:
            print(f"❌ Error leyendo Excel: {e}")
            return
        
        if df.empty:
            print("⚠️  Excel vacío")
            return
        
        # Aplicar filtros
        original_count = len(df)
        
        if blog_filter:
            df = df[df['blog_nombre'] == blog_filter]
            print(f"🔍 Filtro por blog '{blog_filter}': {len(df)}/{original_count} artículos")
        
        if path_filter:
            df = df[df['ruta_archivo'].str.contains(path_filter, case=False, na=False)]
            print(f"🔍 Filtro por ruta '{path_filter}': {len(df)}/{original_count} artículos")
        
        if df.empty:
            print("⚠️  No hay artículos después de aplicar filtros")
            return
        
        print(f"\n{'='*70}")
        print(f"{'🔍 MODO SIMULACION' if dry_run else '✅ ACTUALIZACION REAL'}")
        print(f"📊 Artículos a procesar: {len(df)}")
        print(f"{'='*70}\n")
        
        total_updated = 0
        total_skipped = 0
        total_errors = 0
        
        for idx, row in df.iterrows():
            ruta_archivo = row.get('ruta_archivo')
            if pd.isna(ruta_archivo):
                continue
            
            file_path = self.base_path / ruta_archivo
            
            if not file_path.exists():
                print(f"❌ Archivo no encontrado: {ruta_archivo}")
                total_errors += 1
                continue
            
            try:
                result = self._update_single_qmd(file_path, row, dry_run, idx + 1, len(df))
                if result:
                    total_updated += 1
                else:
                    total_skipped += 1
            except Exception as e:
                print(f"❌ Error en {ruta_archivo}: {e}")
                total_errors += 1
        
        print(f"\n{'='*70}")
        print(f"{'🔍 RESUMEN DE SIMULACION' if dry_run else '✅ RESUMEN DE ACTUALIZACION'}")
        print(f"{'='*70}")
        print(f"✅ Actualizados: {total_updated}")
        print(f"⏭️  Sin cambios: {total_skipped}")
        print(f"❌ Errores: {total_errors}")
        print(f"{'='*70}\n")
        
        if dry_run and total_updated > 0:
            print("💡 Para aplicar cambios, ejecute sin --dry-run\n")
    
    def _update_single_qmd(self, file_path: Path, row: pd.Series, 
                          dry_run: bool, current: int, total: int) -> bool:
        """Actualiza un archivo QMD individual con datos del Excel"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        yaml_match = re.match(r'^---\s*\n(.*?)\n---', content, re.DOTALL)
        if not yaml_match:
            return False
        
        yaml_content = yaml_match.group(1)
        yaml_data = yaml.safe_load(yaml_content) or {}
        
        changes = []
        updated_yaml = self._apply_excel_row_to_yaml(yaml_data, row, changes)
        
        if not changes:
            print(f"[{current}/{total}] ⏭️  Sin cambios: {file_path.parent.name}/{file_path.name}")
            return False
        
        icon = "🔍" if dry_run else "✅"
        action = "Simulando" if dry_run else "Actualizando"
        
        print(f"\n[{current}/{total}] {icon} {action}: {file_path.parent.name}/{file_path.name}")
        print(f"   📝 Cambios detectados: {len(changes)}")
        
        for i, change in enumerate(changes[:5], 1):
            print(f"      {i}. {change}")
        
        if len(changes) > 5:
            print(f"      ... y {len(changes) - 5} cambios más")
        
        if dry_run:
            return True
        
        new_yaml_str = yaml.dump(
            updated_yaml, 
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
            indent=2,
            width=80
        )
        
        new_content = f"---\n{new_yaml_str}---{content[yaml_match.end():]}"
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        return True
    
    def _apply_excel_row_to_yaml(self, yaml_data: Dict, row: pd.Series, 
                                changes: List[str]) -> Dict:
        """Aplica cambios desde una fila de Excel al YAML"""
        
        simple_fields = {
            'title': 'title', 'shorttitle': 'shorttitle', 'subtitle': 'subtitle',
            'date': 'date', 'draft': 'draft', 'abstract': 'abstract',
            'description': 'description', 'image': 'image', 'eval': 'eval',
            'bibliography': 'bibliography', 'course': 'course', 'professor': 'professor',
            'duedate': 'duedate', 'note': 'note', 'journal': 'journal',
            'volume': 'volume', 'copyrightnotice': 'copyrightnotice',
            'copyrightext': 'copyrightext', 'floatsintext': 'floatsintext',
            'mask': 'mask'
        }
        
        for excel_field, yaml_field in simple_fields.items():
            if excel_field in row and not pd.isna(row[excel_field]):
                new_value = row[excel_field]
                old_value = yaml_data.get(yaml_field)
                
                if isinstance(new_value, str) and new_value.upper() in ['TRUE', 'FALSE']:
                    new_value = new_value.upper() == 'TRUE'
                
                # FIX: Convertir copyrightnotice a int
                if excel_field == 'copyrightnotice':
                    try:
                        new_value = int(float(new_value))
                    except:
                        pass
                
                if old_value != new_value:
                    yaml_data[yaml_field] = new_value
                    changes.append(f"{yaml_field}: '{old_value}' → '{new_value}'")
        
        # Campos con guiones
        if 'numbered_lines' in row and not pd.isna(row['numbered_lines']):
            new_value = row['numbered_lines']
            if isinstance(new_value, str):
                new_value = new_value.upper() == 'TRUE'
            old_value = yaml_data.get('numbered-lines')
            if old_value != new_value:
                yaml_data['numbered-lines'] = new_value
                changes.append(f"numbered-lines: {old_value} → {new_value}")
        
        if 'meta_analysis' in row and not pd.isna(row['meta_analysis']):
            new_value = row['meta_analysis']
            if isinstance(new_value, str):
                new_value = new_value.upper() == 'TRUE'
            old_value = yaml_data.get('meta-analysis')
            if old_value != new_value:
                yaml_data['meta-analysis'] = new_value
                changes.append(f"meta-analysis: {old_value} → {new_value}")
        
        # FIX: Tipo de documento - Crea documentmode si cambia de jou
        if 'tipo_documento' in row and not pd.isna(row['tipo_documento']):
            new_type = row['tipo_documento'].lower()
            if new_type in ['stu', 'man', 'jou', 'doc']:
                old_type_direct = yaml_data.get('documentmode')
                old_type_format = None
                
                if 'format' in yaml_data and isinstance(yaml_data['format'], dict):
                    if 'apaquarto-pdf' in yaml_data['format']:
                        apa_config = yaml_data['format']['apaquarto-pdf']
                        if isinstance(apa_config, dict):
                            old_type_format = apa_config.get('documentmode')
                
                old_type = old_type_direct or old_type_format or 'jou'
                
                if old_type != new_type:
                    if 'documentmode' in yaml_data:
                        yaml_data['documentmode'] = new_type
                        changes.append(f"documentmode: {old_type} → {new_type}")
                    elif old_type_format:
                        yaml_data['format']['apaquarto-pdf']['documentmode'] = new_type
                        changes.append(f"format.apaquarto-pdf.documentmode: {old_type} → {new_type}")
                    elif new_type != 'jou':
                        yaml_data['documentmode'] = new_type
                        changes.append(f"documentmode: (jou) → {new_type}")
        
        for field in ['keywords', 'tags', 'categories']:
            if field in row and not pd.isna(row[field]):
                if isinstance(row[field], str):
                    new_value = [item.strip() for item in row[field].split(',') if item.strip()]
                else:
                    new_value = [str(row[field])]
                
                old_value = yaml_data.get(field, [])
                
                # Comparar listas
                if set(old_value) != set(new_value):
                    yaml_data[field] = new_value
                    changes.append(f"{field}: actualizado ({len(new_value)} items)")
        
        # Citation
        if 'citation_type' in row and not pd.isna(row['citation_type']):
            if 'citation' not in yaml_data or not isinstance(yaml_data['citation'], dict):
                yaml_data['citation'] = {}
            new_type = row['citation_type']
            old_type = yaml_data['citation'].get('type')
            if old_type != new_type:
                yaml_data['citation']['type'] = new_type
                changes.append(f"citation.type: {old_type} → {new_type}")
        
        if 'citation_pdf_url' in row and not pd.isna(row['citation_pdf_url']):
            if 'citation' not in yaml_data or not isinstance(yaml_data['citation'], dict):
                yaml_data['citation'] = {}
            new_url = row['citation_pdf_url']
            old_url = yaml_data['citation'].get('pdf-url')
            if old_url != new_url:
                yaml_data['citation']['pdf-url'] = new_url
                changes.append(f"citation.pdf-url actualizada")
        
        # FIX: Autores con corresponding como bool
        authors_data = []
        has_author_in_row = False
        
        for i in range(1, 4):
            prefix = f'author_{i}_'
            if f'{prefix}name' in row and not pd.isna(row[f'{prefix}name']):
                has_author_in_row = True
                author = {'name': row[f'{prefix}name']}
                
                # FIX: Corresponding como bool
                if f'{prefix}corresponding' in row and not pd.isna(row[f'{prefix}corresponding']):
                    corr_val = row[f'{prefix}corresponding']
                    if isinstance(corr_val, str):
                        corr_val = corr_val.upper() == 'TRUE'
                    elif isinstance(corr_val, (int, float)):
                        corr_val = bool(int(corr_val))
                    author['corresponding'] = corr_val
                
                if f'{prefix}orcid' in row and not pd.isna(row[f'{prefix}orcid']):
                    author['orcid'] = row[f'{prefix}orcid']
                
                if f'{prefix}email' in row and not pd.isna(row[f'{prefix}email']):
                    author['email'] = row[f'{prefix}email']
                
                aff = {}
                for aff_field in ['name', 'department', 'city', 'region', 'country']:
                    full_field = f'{prefix}affiliation_{aff_field}'
                    if full_field in row and not pd.isna(row[full_field]):
                        aff[aff_field] = row[full_field]
                
                if aff:
                    author['affiliations'] = [aff]
                
                if f'{prefix}roles' in row and not pd.isna(row[f'{prefix}roles']):
                    roles_str = row[f'{prefix}roles']
                    if isinstance(roles_str, str):
                        author['role'] = [r.strip() for r in roles_str.split(',')]
                
                authors_data.append(author)
        
        # Solo actualizar authors si:
        # 1. Hay autores en el Excel Y
        # 2. Ya existen autores en el index.qmd
        if authors_data and 'author' in yaml_data:
            old_authors = yaml_data.get('author', [])
            if old_authors != authors_data:
                yaml_data['author'] = authors_data
                changes.append(f"author: actualizado ({len(authors_data)} autores)")
        elif authors_data and has_author_in_row:
            # Si no hay author en yaml_data pero sí en Excel, informar pero NO agregar
            # (significa que usa _metadata.yml)
            pass
        
        return yaml_data


def create_config_file(config_path: str, base_path: str):
    """Crea archivo de configuración con valores por defecto"""
    config = {
        'allowed_blogs': [
            'axiomata', 'aequilibria', 'numerus-scriptum',
            'actus-mercator', 'res-publica', 'website-achalma',
            'pecunia-fluxus', 'optimums', 'epsilon-y-beta',
            'methodica', 'chaska', 'dialectica-y-mercado',
        ],
        'excluded_folders': [
            'apa', 'notas', 'borradores',
            'propuesta bicentenario',
            'taller unsch como elaborar tesis de pregrado',
            'practicas preprofesionales',
        ],
        'excel_output_dir': '~/Documents/scripts/scripts_for_quarto/script_metadata_manager/excel_databases'
    }
    
    with open(config_path, 'w', encoding='utf-8') as f:
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False)
    
    print(f"✅ Configuración creada: {config_path}")


def main():
    """Función principal con CLI completo"""
    parser = argparse.ArgumentParser(
        description='Sistema de Gestión de Metadatos Quarto v1.2.1',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:

  # Crear configuración
  python quarto_metadata_manager.py create-config ~/Documents/publicaciones
  
  # Crear base de datos general
  python quarto_metadata_manager.py create-template ~/Documents/publicaciones \\
      --config metadata_config.yml
  
  # Agregar solo articulos nuevos (modo incremental)
  python quarto_metadata_manager.py create-template ~/Documents/publicaciones \\
      --config metadata_config.yml --incremental

  # Simular actualización
  python quarto_metadata_manager.py update ~/Documents/publicaciones \\
      excel_databases/quarto_metadata.xlsx --dry-run
  
  # Actualizar
  python quarto_metadata_manager.py update ~/Documents/publicaciones \\
      excel_databases/quarto_metadata.xlsx
  
  # Actualizar solo un blog
  python quarto_metadata_manager.py update ~/Documents/publicaciones \\
      excel_databases/quarto_metadata.xlsx --blog axiomata
  
  # Actualizar solo rutas específicas
  python quarto_metadata_manager.py update ~/Documents/publicaciones \\
      excel_databases/quarto_metadata.xlsx --filter-path "posts/2024"

Versión: 1.2.0
Autor: Edison Achalma
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Comandos')
    
    # create-config
    parser_config = subparsers.add_parser('create-config')
    parser_config.add_argument('base_path')
    parser_config.add_argument('-o', '--output', default='metadata_config.yml')
    
    # create-template
    parser_create = subparsers.add_parser('create-template')
    parser_create.add_argument('base_path')
    parser_create.add_argument('-o', '--output', default='quarto_metadata.xlsx')
    parser_create.add_argument('-b', '--blog')
    parser_create.add_argument('-c', '--config')
    parser_create.add_argument('--incremental', action='store_true',
                              help='Solo agregar nuevos articulos')
    
    # update
    parser_update = subparsers.add_parser('update')
    parser_update.add_argument('base_path')
    parser_update.add_argument('excel_file')
    parser_update.add_argument('-b', '--blog', help='Filtrar por blog')
    parser_update.add_argument('-p', '--filter-path', help='Filtrar por ruta')
    parser_update.add_argument('-c', '--config')
    parser_update.add_argument('--dry-run', action='store_true')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    try:
        if args.command == 'create-config':
            create_config_file(args.output, args.base_path)
            
        elif args.command == 'create-template':
            config_file = getattr(args, 'config', None)
            manager = QuartoMetadataManager(args.base_path, config_file)
            
            output_path = args.output
            if args.blog:
                name, ext = os.path.splitext(output_path)
                output_path = f"{name}_{args.blog}{ext}"
            
            incremental = getattr(args, 'incremental', False)
            manager.create_excel_template(output_path, args.blog, incremental)
            
        elif args.command == 'update':
            config_file = getattr(args, 'config', None)
            manager = QuartoMetadataManager(args.base_path, config_file)
            
            blog_filter = getattr(args, 'blog', None)
            path_filter = getattr(args, 'filter_path', None)
            
            manager.update_yaml_from_excel(
                args.excel_file, 
                blog_filter, 
                path_filter,
                args.dry_run
            )
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == '__main__':
    sys.exit(main())